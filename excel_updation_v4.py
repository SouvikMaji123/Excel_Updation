import re
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def remove_blank_rows(df):
    return df.dropna(how='all')

def merge_columns(df, columns, new_col_name):
    df[new_col_name] = df[columns].astype(str).agg('-'.join, axis=1)
    return df

def update_excel_sync(test_df, final_data):
    try:
        wb = load_workbook(final_data)
        ws = wb.active
        
        pattern = re.compile(r'~?input needed~?', re.IGNORECASE)
        
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if pattern.match(str(cell_value).strip()) and pd.notna(test_df.iat[row-2, col-1]):
                    ws.cell(row=row, column=col).value = test_df.iat[row-2, col-1]
        
        updated_final_file = "updated_final_excel.xlsx"
        wb.save(updated_final_file)
        return updated_final_file
    except Exception as e:
        st.error(f"Error updating Excel file: {e}")
        return None

def update_excel_not_sync(test_df, final_data, id_column):
    try:
        wb = load_workbook(final_data)
        ws = wb.active
        
        test_dict = test_df.set_index(id_column).to_dict('index')
        pattern = re.compile(r'~?input needed~?', re.IGNORECASE)
             
        for row in range(2, ws.max_row + 1):
            row_id = ws.cell(row=row, column=1).value
            if row_id in test_dict:
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if pattern.match(str(cell_value).strip()):
                        col_name = ws.cell(row=1, column=col).value.lower()
                        if col_name in test_dict[row_id] and pd.notna(test_dict[row_id][col_name]):
                            ws.cell(row=row, column=col).value = test_dict[row_id][col_name]
        
        updated_final_file = "updated_final_excel.xlsx"
        wb.save(updated_final_file)
        return updated_final_file
    except Exception as e:
        st.error(f"Error updating Excel file: {e}")
        return None

def update_excel_with_merged_columns(test_df, final_data, columns_to_merge):
    try:
        wb = load_workbook(final_data)
        ws = wb.active
        
        # Merge specified columns in test_df
        test_df = merge_columns(test_df, columns_to_merge, 'merged_column')
        
        # Merge specified columns in final Excel
        final_df = pd.DataFrame(ws.values)
        final_df.columns = final_df.iloc[0].str.lower()
        final_df = final_df[1:]
        final_df = merge_columns(final_df, columns_to_merge, 'merged_column')
        
        test_dict = test_df.set_index('merged_column').to_dict('index')
        pattern = re.compile(r'~?input needed~?', re.IGNORECASE)
        
        for row in range(2, ws.max_row + 1):
            merged_value = '-'.join([str(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value.lower() in columns_to_merge])
            if merged_value in test_dict:
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if pattern.match(str(cell_value).strip()):
                        col_name = ws.cell(row=1, column=col).value.lower()
                        if col_name in test_dict[merged_value] and pd.notna(test_dict[merged_value][col_name]):
                            ws.cell(row=row, column=col).value = test_dict[merged_value][col_name]
        
        
        updated_final_file = "updated_final_excel.xlsx"
        wb.save(updated_final_file)
        return updated_final_file
    except Exception as e:
        st.error(f"Error updating Excel file: {e}")
        return None

def filter_columns_with_pattern(df, pattern):
    columns_to_exclude = []
    for col in df.columns:
        if df[col].astype(str).str.contains(pattern, flags=re.IGNORECASE, na=False).any():
            columns_to_exclude.append(col)
    return [col for col in df.columns if col not in columns_to_exclude]

st.sidebar.title("Excel Updater")

# Initialize updated_final_file
updated_final_file = None

# Upload test Excel file
test_file = st.sidebar.file_uploader("Upload Test Excel", type=["xlsx"])
# Upload final Excel file
final_file = st.sidebar.file_uploader("Upload Final Excel", type=["xlsx"])

if test_file and final_file:
    test_df = pd.read_excel(test_file)
    test_df.columns = map(str.lower, test_df.columns)  # Convert DataFrame column names to lowercase
    test_df = remove_blank_rows(test_df)  # Remove blank rows from test_df
    
    final_df = pd.read_excel(final_file)
    final_df.columns = map(str.lower, final_df.columns)  # Convert DataFrame column names to lowercase

    # Radio buttons to ask if rows are in sync
    rows_in_sync = st.radio("Are the rows in sync between the test and final Excel files?", ("Yes", "No"))
    
    if rows_in_sync == "Yes":
        if len(test_df) != len(final_df): # Check if the number of rows are the same
            st.warning("The number of rows in the test and final Excel files are different. Please select 'No' for more options .")
        else:
            updated_final_file = update_excel_sync(test_df, final_file)
 
    else:
        st.write("Test DataFrame Columns:", test_df.columns.tolist())
    
        unique_column_present = st.radio("Is there a unique column present in both Excel files?", ("Yes", "No"))
    
        if unique_column_present == "Yes":
            pattern = r'~?input needed~?'
            filtered_columns = filter_columns_with_pattern(final_df, pattern)
            id_column = st.selectbox("Select the unique column present in both Excel files", filtered_columns, key="id_input")
            if id_column and st.button("Submit"):
                if id_column not in test_df.columns:
                    st.error(f"Column '{id_column}' not found in both Excel files. Please enter a valid column name.")
                else:
                    updated_final_file = update_excel_not_sync(test_df, final_file, id_column)
        else:
            pattern = r'~?input needed~?'
            filtered_columns = filter_columns_with_pattern(final_df, pattern)
            columns_to_combine = st.multiselect("Select columns to combine to create a unique identifier", filtered_columns)
            if all(columns_to_combine) and st.button("Submit"):
                if any(col not in test_df.columns for col in columns_to_combine):
                    st.error("One or more columns specified are not found in both Excel files. Please enter valid column names.")
                else:
                    # Debugging: Print combined columns
                    st.write("Combining columns:", columns_to_combine)
                    updated_final_file = update_excel_with_merged_columns(test_df, final_file, columns_to_combine)

    if updated_final_file:
        # Provide a download link for the updated final Excel
        with open(updated_final_file, "rb") as file:
            btn = st.download_button(
                label="Download Updated Final Excel",
                data=file,
                file_name=updated_final_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
